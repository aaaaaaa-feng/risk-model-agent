from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from app.core.security import sha256_file
from app.workers.model_package import build_model_package
from app.workers.modeling import ModelBundle, ScorecardEstimator
from app.workers.package_runtime import (
    inspect_skops_types,
    safe_extract_model_package,
    score_package_directory,
    verify_package_directory,
)


class UnapprovedSkopsType:
    def __init__(self):
        self.value = "must-not-load"


def test_full_agent_reviewer_worker_pipeline(golden):
    context = golden["context"]
    run = context.catalog.require("runs", golden["run"]["id"])
    state = run["state"]
    assert run["status"] == "succeeded"
    assert state["split"]["customer_key"] == "customer_id"
    frame = context.catalog.dataset_frame(state["working_dataset_version_id"])
    partitions = state["split"]["indices"]
    customer_sets = {
        name: set(frame.iloc[indices]["customer_id"]) for name, indices in partitions.items()
    }
    assert customer_sets["train"].isdisjoint(customer_sets["test"])
    assert customer_sets["train"].isdisjoint(customer_sets["oot"])
    assert customer_sets["test"].isdisjoint(customer_sets["oot"])
    assert state["model_result"]["oot_used_for_selection"] is False
    statuses = {item["candidate"]: item["status"] for item in state["model_result"]["candidates"]}
    assert statuses == {
        "dummy": "trained",
        "scorecard": "trained",
        "regularized_logistic": "trained",
        "xgboost": "trained",
    }
    excluded = {item["column"]: item["reason"] for item in state["screening"]["excluded"]}
    assert excluded["FPD7"] == "OTHER_TARGET"
    assert excluded["MOB30"] == "OTHER_TARGET"
    reviews = context.database.list("review_records", {"run_id": run["id"]}, limit=200)
    assert {"code", "execution", "report"}.issubset({item["scope"] for item in reviews})
    assert all(item["status"] == "fallback_pass" for item in reviews)
    assert context.engine.persistence_mode == "sqlite"


def test_report_excel_html_json_share_one_structured_source(golden):
    context = golden["context"]
    run = context.catalog.require("runs", golden["run"]["id"])
    artifacts = context.database.list("artifacts", {"run_id": run["id"]}, limit=100)
    by_kind = {item["kind"]: item for item in artifacts}
    assert {
        "report_json",
        "report_excel",
        "report_html",
        "model_package",
        "reproducible_notebook",
    }.issubset(by_kind)
    for artifact in artifacts:
        assert Path(artifact["path"]).exists()
        assert sha256_file(Path(artifact["path"])) == artifact["checksum"]
    report = json.loads(Path(by_kind["report_json"]["path"]).read_text(encoding="utf-8"))
    html = Path(by_kind["report_html"]["path"]).read_text(encoding="utf-8")
    assert report["schema_version"] == "risk-model-report/v1"
    coverage = report["review"]["coverage"]
    assert coverage["total_records"] == len(report["review"]["records"])
    assert coverage["deterministic_coverage"] == 1.0
    assert coverage["llm_coverage"] == 0.0
    assert coverage["fallback_rate"] > 0
    assert report["executive_summary"]["champion"] == report["champion"]["candidate"]
    expected_verdict = "pass" if report["executive_summary"]["absolute_ordering"] else "conditional"
    assert report["executive_summary"]["quality_verdict"] == expected_verdict
    if expected_verdict == "conditional":
        assert report["executive_summary"]["quality_notes"]
    assert 'id="risk-model-report-data"' in html
    assert json.dumps(report["executive_summary"]["champion"], ensure_ascii=False) in html
    workbook = openpyxl.load_workbook(
        by_kind["report_excel"]["path"], read_only=True, data_only=True
    )
    assert workbook.sheetnames[:3] == ["总体情况", "入模变量", "模型分箱"]
    assert workbook["总体情况"]["C3"].value == report["executive_summary"]["champion"]
    feature_rows = list(workbook["入模变量"].iter_rows(values_only=True))
    flattened = {str(value) for row in feature_rows for value in row if value is not None}
    assert {item["column"] for item in report["feature_selection"]["selected"]}.issubset(flattened)


def test_persisted_model_reload_scores_identically_and_names_columns(golden):
    context = golden["context"]
    run = context.catalog.require("runs", golden["run"]["id"])
    model = context.database.list("model_versions", {"run_id": run["id"]}, limit=10)[0]
    frame = context.catalog.dataset_frame(golden["demo"]["dataset_version"]["id"])
    source = Path(golden["demo"]["dataset_version"]["stored_path"])
    asset = context.catalog.register_asset(
        golden["demo"]["project"]["id"], source, "score_input.csv", "score_input"
    )
    first_job, first_artifact = context.artifacts.score_file(model["id"], asset["id"])
    second_job, second_artifact = context.artifacts.score_file(model["id"], asset["id"])
    import pandas as pd

    first = pd.read_csv(first_artifact["path"])
    second = pd.read_csv(second_artifact["path"])
    score_column = first_job["metadata"]["score_column"]
    assert score_column.startswith(model["name"].replace("-", "_"))
    assert np.array_equal(first[score_column].to_numpy(), second[score_column].to_numpy())
    assert np.allclose(
        first[f"{score_column}_bad_probability"],
        second[f"{score_column}_bad_probability"],
    )
    assert first[score_column].between(300, 900).all()
    assert second_job["rows"] == len(frame)


def test_model_package_contract_and_hashes(golden):
    context = golden["context"]
    run = context.catalog.require("runs", golden["run"]["id"])
    state = run["state"]
    manifest = state["package_manifest"]
    assert manifest["raw_data_included"] is False
    champion = state["model_result"]["champion"]
    assert manifest["schema_version"] == "risk-model-package/v2"
    if champion == "scorecard":
        assert any(item["format"] == "scorecard-json" for item in manifest["formats"])
        assert not any(item["format"] == "skops" for item in manifest["formats"])
    else:
        assert any(item["format"] == "skops" for item in manifest["formats"])
    expected_native = {
        "xgboost": "xgboost-json",
        "lightgbm": "lightgbm-text",
        "catboost": "catboost-cbm",
    }.get(champion)
    if expected_native:
        assert any(item["format"] == expected_native for item in manifest["formats"])
    package = next(
        item
        for item in context.database.list("artifacts", {"run_id": run["id"]}, limit=100)
        if item["kind"] == "model_package"
    )
    assert manifest["package_sha256"] == sha256_file(Path(package["path"]))


def test_package_hashes_are_verified_before_loading_and_cli_scores_cleanly(golden, tmp_path: Path):
    context = golden["context"]
    run = context.catalog.require("runs", golden["run"]["id"])
    package_record = next(
        item
        for item in context.database.list("artifacts", {"run_id": run["id"]}, limit=100)
        if item["kind"] == "model_package"
    )
    package = Path(package_record["path"])
    root = safe_extract_model_package(package, tmp_path / "package")
    verify_package_directory(root)
    contract = json.loads((root / "field_contract.json").read_text(encoding="utf-8"))
    source = context.catalog.dataset_frame(golden["demo"]["dataset_version"]["id"])
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    source[contract["required_fields"]].head(25).to_csv(input_path, index=False)
    environment = os.environ.copy()
    environment.pop("PYTHONPATH", None)
    environment["PYTHONNOUSERSITE"] = "1"
    completed = subprocess.run(
        [
            sys.executable,
            str(root / "score.py"),
            "--input",
            str(input_path),
            "--output",
            str(output_path),
        ],
        cwd=tmp_path,
        env=environment,
        text=True,
        capture_output=True,
        timeout=120,
        check=False,
    )
    assert completed.returncode == 0, completed.stderr
    assert len(pd.read_csv(output_path)) == 25
    assert "from app" not in (root / "risk_model_agent_package_runtime.py").read_text(
        encoding="utf-8"
    )

    manifest_path = root / "manifest.json"
    original_manifest = manifest_path.read_text(encoding="utf-8")
    tampered_manifest = json.loads(original_manifest)
    tampered_manifest["formats"][0]["name"] = "../outside-model"
    manifest_path.write_text(json.dumps(tampered_manifest), encoding="utf-8")
    with pytest.raises(ValueError, match="MODEL_PACKAGE_FILE_REFERENCE_INVALID"):
        verify_package_directory(root)
    manifest_path.write_text(original_manifest, encoding="utf-8")

    (root / "field_contract.json").write_text("{}", encoding="utf-8")
    with pytest.raises(ValueError, match="MODEL_PACKAGE_FILE_HASH_MISMATCH"):
        verify_package_directory(root)


def test_scorecard_package_uses_json_rules_and_matches_estimator(tmp_path: Path):
    rng = np.random.default_rng(42)
    frame = pd.DataFrame(
        {
            "income": rng.normal(8_000, 1_500, 240),
            "segment": rng.choice(["A", "B", "C"], 240),
        }
    )
    target = ((frame["income"] < 7_500) | (frame["segment"] == "C")).astype(int)
    estimator = ScorecardEstimator().fit(frame, target)
    bundle = ModelBundle(
        "scorecard-test",
        "scorecard",
        estimator,
        ["income", "segment"],
        "uncalibrated",
        {"minimum": 300, "maximum": 900, "base_score": 600, "base_odds": 20, "pdo": 50},
        {},
    )
    contract = {
        "schema_version": "risk-field-contract/v2",
        "model_name": bundle.name,
        "required_fields": bundle.features,
        "dtypes": {column: str(frame[column].dtype) for column in bundle.features},
        "field_types": {"income": "numeric", "segment": "categorical"},
    }
    package, manifest = build_model_package(bundle, contract, tmp_path / "scorecard.zip", [])
    assert any(item["format"] == "scorecard-json" for item in manifest["formats"])
    assert not any(item["format"] == "skops" for item in manifest["formats"])
    root = safe_extract_model_package(package, tmp_path / "scorecard")
    probability, _, _, _ = score_package_directory(root, frame)
    assert np.allclose(probability, estimator.predict_proba(frame)[:, 1])


def test_skops_type_policy_rejects_unknown_classes(tmp_path: Path):
    import skops.io as sio

    path = tmp_path / "unapproved.skops"
    sio.dump(UnapprovedSkopsType(), path)
    with pytest.raises(ValueError, match="MODEL_SKOPS_TYPE_NOT_APPROVED"):
        inspect_skops_types(path, "regularized_logistic")


def test_worker_bundle_manifest_is_bound_to_checkpoint_state(golden):
    context = golden["context"]
    run = context.catalog.require("runs", golden["run"]["id"])
    state = run["state"]
    manifest = context.pipeline._bundle_dir(run["id"]) / "manifest.json"
    original = manifest.read_bytes()
    try:
        manifest.write_text("{}", encoding="utf-8")
        with pytest.raises(ValueError, match="WORKER_BUNDLE_MANIFEST_CHECKSUM_MISMATCH"):
            context.pipeline._load_bundles(run["id"], state["worker_bundle_manifest_sha256"])
    finally:
        manifest.write_bytes(original)


def test_full_golden_project_archive_restores_self_contained_paths(golden):
    context = golden["context"]
    project_id = golden["demo"]["project"]["id"]
    archive, recovery_key = context.archives.create(project_id, "golden-archive-password")
    restored = context.archives.restore(Path(archive["path"]), recovery_key)
    restored_datasets = context.database.list_all(
        "dataset_versions", {"project_id": restored["id"]}
    )
    restored_runs = context.database.list_all("runs", {"project_id": restored["id"]})
    assert len(restored_datasets) == len(
        context.database.list_all("dataset_versions", {"project_id": project_id})
    )
    assert len(restored_runs) == len(context.database.list_all("runs", {"project_id": project_id}))
    for table in ("data_assets", "dataset_versions"):
        for row in context.database.list_all(table, {"project_id": restored["id"]}):
            path = Path(row["stored_path"])
            assert path.is_file()
            assert context.paths.project_dir(restored["id"]) in path.parents
    for run in restored_runs:
        for artifact in context.database.list_all("artifacts", {"run_id": run["id"]}):
            path = Path(artifact["path"])
            assert path.is_file()
            assert sha256_file(path) == artifact["checksum"]
