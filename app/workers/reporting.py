from __future__ import annotations

import html
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.security import sha256_file


REPORT_SCHEMA = "risk-model-report/v1"
HEADER_FILL = PatternFill("solid", fgColor="16324F")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCEAF7")


def build_report(
    *,
    project: dict[str, Any],
    run: dict[str, Any],
    target_task: dict[str, Any],
    dataset: dict[str, Any],
    diagnostics: dict[str, Any],
    split: dict[str, Any],
    screening: dict[str, Any],
    binning: dict[str, Any],
    model_result: dict[str, Any],
    bin_reports: dict[str, list[dict[str, Any]]],
    reviews: list[dict[str, Any]],
    lineage: dict[str, Any],
) -> dict[str, Any]:
    champion_name = model_result["champion"]
    champion = next(
        item for item in model_result["candidates"] if item["candidate"] == champion_name
    )
    samples = {
        name: {
            "rows": int(details.get("rows") or 0),
            "positive_count": int(details.get("positive_count") or 0),
            "negative_count": int(details.get("negative_count") or 0),
            "bad_rate": details.get("bad_rate"),
        }
        for name, details in model_result["champion_metrics"].items()
        if details
    }
    selected = [
        item
        for item in screening.get("features", [])
        if item.get("status") == "included"
    ]
    absolute_ordering = bool(
        (champion.get("test_monotonicity") or {}).get("absolute")
    )
    review_passed = bool(reviews and reviews[-1].get("status") == "pass")
    quality_verdict = "pass" if review_passed and absolute_ordering else "conditional"
    quality_notes: list[str] = []
    if not absolute_ordering:
        quality_notes.append(
            "Test 等频分箱未达到绝对排序；模型可用于分析和后续调整，但不能标记为无条件通过。"
        )
    if not review_passed:
        quality_notes.append("最近一轮 Reviewer 记录不是 pass，请复核质检证据。")
    report = {
        "schema_version": REPORT_SCHEMA,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "project": {"id": project["id"], "name": project["name"]},
        "run": {
            "id": run["id"],
            "mode": run["mode"],
            "status": run["status"],
            "stage": run["stage"],
        },
        "target": {
            "task_id": target_task["id"],
            "column": target_task["target_column"],
            "labels": target_task.get("labels", {}),
            "valid_sample_count": target_task.get("valid_sample_count", 0),
        },
        "dataset": {
            "version_id": dataset["id"],
            "label": dataset["label"],
            "rows": dataset["rows"],
            "columns": dataset["columns"],
            "lineage": lineage,
        },
        "executive_summary": {
            "champion": champion_name,
            "test_auc": champion["test_metrics"].get("roc_auc"),
            "test_ks": champion["test_metrics"].get("ks"),
            "oot_auc": (champion.get("oot_metrics") or {}).get("roc_auc"),
            "oot_ks": (champion.get("oot_metrics") or {}).get("ks"),
            "absolute_ordering": absolute_ordering,
            "selected_feature_count": len(selected),
            "quality_verdict": quality_verdict,
            "quality_notes": quality_notes,
        },
        "sample_overview": samples,
        "diagnostics": diagnostics,
        "split": {
            key: value
            for key, value in split.items()
            if key not in {"indices", "row_ids"}
        },
        "feature_selection": {
            "thresholds": screening.get("thresholds", {}),
            "fit_scope": screening.get("fit_scope"),
            "selected": selected,
            "excluded": screening.get("excluded", []),
            "correlation_pairs": screening.get("correlation_pairs", []),
            "restored": screening.get("restored", []),
        },
        "binning": {
            "version": binning.get("version"),
            "fit_scope": binning.get("fit_scope"),
            "specs": binning.get("specs", {}),
            "datasets": bin_reports,
        },
        "model_comparison": model_result["candidates"],
        "champion": champion,
        "score": model_result["score"],
        "review": {
            "records": reviews,
            "max_repair_rounds": 3,
            "independent_context": True,
        },
        "governance": {
            "raw_data_uploaded": False,
            "train_only_fit": True,
            "oot_used_for_selection": False,
            "small_cell_suppression_threshold": 30,
            "hidden_chain_of_thought_included": False,
        },
    }
    return json.loads(json.dumps(report, ensure_ascii=False, default=_json_default))


def write_report_json(report: dict[str, Any], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _json_default(value: Any) -> Any:
    if hasattr(value, "item"):
        return value.item()
    if hasattr(value, "tolist"):
        return value.tolist()
    return str(value)


def _flatten_metrics(report: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for partition, values in report.get("sample_overview", {}).items():
        metrics = (report.get("champion", {}).get(f"{partition}_metrics") or {})
        rows.append(
            [
                partition.upper(),
                values.get("rows"),
                values.get("positive_count"),
                values.get("negative_count"),
                values.get("bad_rate"),
                metrics.get("roc_auc"),
                metrics.get("ks"),
                metrics.get("pr_auc"),
                report.get("champion", {}).get(
                    "train_test_score_psi" if partition == "test" else "test_oot_score_psi"
                ),
            ]
        )
    return rows


def _append_table(sheet: Any, title: str, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    sheet.append([title])
    sheet.cell(sheet.max_row, 1).font = Font(bold=True, size=14, color="16324F")
    sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=len(headers))
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    sheet.append([])


def _excel_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _finish_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        width = min(42, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_report_excel(report: dict[str, Any], path: Path) -> Path:
    workbook = Workbook()
    overall = workbook.active
    overall.title = "总体情况"
    summary = report["executive_summary"]
    _append_table(
        overall,
        "模型管理摘要",
        ["项目", "Y", "Champion", "Test AUC", "Test KS", "OOT AUC", "OOT KS", "绝对排序", "质检结论", "质检说明"],
        [[
            report["project"]["name"],
            report["target"]["column"],
            summary["champion"],
            summary["test_auc"],
            summary["test_ks"],
            summary["oot_auc"],
            summary["oot_ks"],
            summary["absolute_ordering"],
            summary["quality_verdict"],
            "；".join(summary.get("quality_notes", [])),
        ]],
    )
    _append_table(
        overall,
        "Train / Test / OOT",
        ["数据集", "样本量", "坏样本", "好样本", "坏占比", "AUC", "KS", "PR-AUC", "PSI"],
        _flatten_metrics(report),
    )
    _finish_sheet(overall)

    features = workbook.create_sheet("入模变量")
    _append_table(
        features,
        "最终入模变量（仅展示最终入模）",
        ["变量", "类型", "缺失率", "IV", "人工恢复", "恢复理由"],
        [
            [
                item.get("column"), item.get("type"), item.get("missing_rate"), item.get("iv"),
                item.get("restored", False), item.get("restore_reason", ""),
            ]
            for item in report["feature_selection"]["selected"]
        ],
    )
    importance = {item["feature"]: item["importance"] for item in report["champion"].get("feature_importance", [])}
    _append_table(
        features,
        "特征重要性",
        ["变量", "重要性"],
        sorted(importance.items(), key=lambda item: item[1], reverse=True),
    )
    bin_rows: list[list[Any]] = []
    for feature, spec in report["binning"]["specs"].items():
        for row in spec.get("table", []):
            bin_rows.append([
                feature, spec.get("source"), spec.get("monotonic"), row.get("bin"),
                row.get("count"), row.get("good"), row.get("bad"), row.get("bad_rate"),
                row.get("woe"), row.get("iv"), spec.get("iv"),
            ])
    _append_table(
        features,
        "最终入模变量分箱（Train 拟合）",
        ["变量", "来源", "单调", "分箱", "样本量", "好样本", "坏样本", "坏占比", "WOE", "箱IV", "变量IV"],
        bin_rows,
    )
    _finish_sheet(features)

    performance = workbook.create_sheet("模型分箱")
    performance_rows: list[list[Any]] = []
    for dataset, rows in report["champion"].get("lift", {}).items():
        for row in rows:
            performance_rows.append([
                dataset.upper(), row.get("bucket"), row.get("count"), row.get("bad"),
                row.get("bad_rate"), row.get("lift"), row.get("cumulative_capture"),
                row.get("min_probability"), row.get("max_probability"),
            ])
    _append_table(
        performance,
        "模型等频分箱",
        ["数据集", "箱", "样本量", "坏样本", "坏占比", "Lift", "累计捕获", "最小坏概率", "最大坏概率"],
        performance_rows,
    )
    _finish_sheet(performance)

    comparison = workbook.create_sheet("模型对比")
    _append_table(
        comparison,
        "候选模型对比",
        ["模型", "校准", "Test AUC", "Test KS", "Test PR-AUC", "Test Brier", "排序性", "选择分"],
        [[
            item["candidate"], item["calibration"], item["test_metrics"].get("roc_auc"),
            item["test_metrics"].get("ks"), item["test_metrics"].get("pr_auc"),
            item["test_metrics"].get("brier"), item["test_monotonicity"].get("absolute"),
            item["selection_score"],
        ] for item in report["model_comparison"]],
    )
    _finish_sheet(comparison)

    review = workbook.create_sheet("质检记录")
    _append_table(
        review,
        "独立 Reviewer 质检",
        ["轮次", "范围", "状态", "问题", "证据"],
        [[item.get("round"), item.get("scope"), item.get("status"), item.get("issues"), item.get("evidence")] for item in report["review"]["records"]],
    )
    _finish_sheet(review)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _metric(value: Any) -> str:
    return "—" if value is None else f"{float(value):.4f}"


def _html_rows(headers: list[str], rows: Iterable[Iterable[Any]]) -> str:
    head = "".join(f"<th>{html.escape(str(value))}</th>" for value in headers)
    body = "".join(
        "<tr>" + "".join(f"<td>{html.escape(str(_excel_value(value) if value is not None else '—'))}</td>" for value in row) + "</tr>"
        for row in rows
    )
    return f"<div class=table-wrap><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>"


def write_report_html(report: dict[str, Any], path: Path) -> Path:
    summary = report["executive_summary"]
    comparison_rows = [
        [item["candidate"], item["calibration"], _metric(item["test_metrics"].get("roc_auc")), _metric(item["test_metrics"].get("ks")), item["test_monotonicity"].get("absolute")]
        for item in report["model_comparison"]
    ]
    lift_rows = []
    for dataset, rows in report["champion"].get("lift", {}).items():
        for row in rows:
            lift_rows.append([dataset.upper(), row["bucket"], row["count"], row["bad"], _metric(row["bad_rate"]), _metric(row["lift"]), _metric(row["cumulative_capture"])])
    serialized = html.escape(json.dumps(report, ensure_ascii=False), quote=False)
    quality_note = "；".join(summary.get("quality_notes", [])) or "Reviewer 与确定性检查均已通过。"
    document = f"""<!doctype html>
<html lang=\"zh-CN\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1\">
<title>{html.escape(report['project']['name'])} · 风控模型报告</title><style>
:root{{--ink:#17202a;--muted:#657587;--blue:#176b87;--teal:#1a8f83;--line:#dce5eb;--soft:#f4f7f9;--warn:#c67b19}}
*{{box-sizing:border-box}}body{{margin:0;background:#eef3f5;color:var(--ink);font:14px/1.55 Inter,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif}}
.page{{max-width:1200px;margin:32px auto;background:white;border:1px solid var(--line);box-shadow:0 16px 48px #102a3b16}}
header{{padding:34px 42px;background:linear-gradient(135deg,#102f46,#176b87);color:white}}h1{{margin:0 0 8px;font-size:28px}}header p{{margin:0;opacity:.78}}
main{{padding:28px 42px 50px}}h2{{margin:34px 0 14px;font-size:19px;border-left:4px solid var(--teal);padding-left:10px}}
.cards{{display:grid;grid-template-columns:repeat(5,1fr);gap:12px}}.card{{border:1px solid var(--line);border-radius:10px;padding:14px;background:var(--soft)}}.card span{{display:block;color:var(--muted);font-size:12px}}.card b{{font-size:22px}}
.table-wrap{{overflow:auto;border:1px solid var(--line);border-radius:10px}}table{{width:100%;border-collapse:collapse;white-space:nowrap}}th,td{{padding:10px 12px;border-bottom:1px solid var(--line);text-align:left}}th{{background:#eaf2f6;color:#26465a}}tr:last-child td{{border:0}}
.note{{padding:13px 15px;background:#fff8e9;border:1px solid #f0d7a4;border-radius:8px;color:#765118}}footer{{padding:18px 42px;border-top:1px solid var(--line);color:var(--muted)}}
@media(max-width:800px){{.page{{margin:0;border:0}}header,main,footer{{padding-left:20px;padding-right:20px}}.cards{{grid-template-columns:repeat(2,1fr)}}}}
@media print{{body{{background:white}}.page{{margin:0;box-shadow:none;border:0}}}}
</style></head><body><article class=page><header><h1>风控模型报告</h1><p>{html.escape(report['project']['name'])} · Y={html.escape(report['target']['column'])} · Run {html.escape(report['run']['id'])}</p></header><main>
<section class=cards><div class=card><span>Champion</span><b>{html.escape(str(summary['champion']))}</b></div><div class=card><span>Test AUC</span><b>{_metric(summary['test_auc'])}</b></div><div class=card><span>Test KS</span><b>{_metric(summary['test_ks'])}</b></div><div class=card><span>OOT AUC</span><b>{_metric(summary['oot_auc'])}</b></div><div class=card><span>最终变量</span><b>{summary['selected_feature_count']}</b></div></section>
<h2>管理摘要</h2><p class=note>质检结论：{html.escape(str(summary['quality_verdict']))}。{html.escape(quality_note)} 本报告中的训练、筛选与分箱仅在 Train 上拟合；Test 用于方案选择，OOT 只用于最终报告。AUC/KS 阈值为参考，不作为硬性业务准入。</p>
<h2>候选模型对比</h2>{_html_rows(['模型','校准','Test AUC','Test KS','绝对排序'], comparison_rows)}
<h2>Train / Test / OOT</h2>{_html_rows(['数据集','样本量','坏样本','好样本','坏占比','AUC','KS','PR-AUC','PSI'], _flatten_metrics(report))}
<h2>模型分箱</h2>{_html_rows(['数据集','箱','样本量','坏样本','坏占比','Lift','累计捕获'], lift_rows)}
<h2>最终入模变量</h2>{_html_rows(['变量','类型','缺失率','IV','人工恢复'], [[item.get('column'),item.get('type'),_metric(item.get('missing_rate')),_metric(item.get('iv')),item.get('restored',False)] for item in report['feature_selection']['selected']])}
<h2>质检结论</h2>{_html_rows(['轮次','范围','状态','问题'], [[item.get('round'),item.get('scope'),item.get('status'),item.get('issues')] for item in report['review']['records']])}
<script type=\"application/json\" id=\"risk-model-report-data\">{serialized}</script></main><footer>Risk Model Agent · 单文件离线报告 · {html.escape(report['generated_at'])}</footer></article></body></html>"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(document, encoding="utf-8")
    return path


def artifact_manifest(paths: Iterable[Path]) -> list[dict[str, Any]]:
    return [
        {"name": path.name, "size_bytes": path.stat().st_size, "sha256": sha256_file(path)}
        for path in paths
    ]
