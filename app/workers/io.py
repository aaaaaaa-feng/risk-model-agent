from __future__ import annotations

import csv
import math
import os
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

from app.core.security import sha256_file


SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xlsm", ".xls"}


@dataclass(frozen=True)
class ResourcePlan:
    available_memory_bytes: int
    budget_bytes: int
    estimated_memory_bytes: int
    row_chunk_size: int
    column_batch_size: int
    max_parallel_models: int
    strategy: str
    warnings: list[str]

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


def available_memory_bytes() -> int:
    try:
        import psutil

        return int(psutil.virtual_memory().available)
    except ImportError:
        pass
    try:
        pages = os.sysconf("SC_AVPHYS_PAGES")
        page_size = os.sysconf("SC_PAGE_SIZE")
        return int(pages * page_size)
    except (AttributeError, OSError, ValueError):
        return 4 * 1024**3


def plan_resources(rows: int, columns: int, memory_budget_mb: int = 1536) -> ResourcePlan:
    available = available_memory_bytes()
    configured = max(256, memory_budget_mb) * 1024**2
    budget = min(configured, max(256 * 1024**2, int(available * 0.45)))
    # 24 bytes/cell is a conservative mixed numeric/string planning estimate;
    # the worker process additionally enforces live RSS and a hard timeout.
    estimated = max(1, rows) * max(1, columns) * 24
    ratio = estimated / max(budget, 1)
    column_batch = max(8, min(columns or 8, int((budget / max(rows, 1) / 24) or 8)))
    column_batch = min(column_batch, 128)
    row_chunk = max(1000, min(rows or 1000, int((budget / max(columns, 1) / 24) or 1000)))
    row_chunk = min(row_chunk, 100_000)
    warnings: list[str] = []
    if ratio <= 0.55:
        strategy = "in_memory"
    elif ratio <= 1.5:
        strategy = "column_batched"
        warnings.append("将按列分批计算画像、缺失率和 IV。")
    else:
        strategy = "requires_projection"
        warnings.append("完整表预计超过 Worker 内存上限，必须先做列投影、提高预算或延后执行。")
        warnings.append("系统会明确阻断，不会静默删样本，也不会用线程继续冒险加载。")
    return ResourcePlan(
        available,
        budget,
        estimated,
        row_chunk,
        column_batch,
        1 if ratio > 0.35 else 2,
        strategy,
        warnings,
    )


def detect_csv_encoding(path: Path) -> str:
    with path.open("rb") as stream:
        sample = stream.read(131072)
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "big5"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "latin1"


def sniff_delimiter(path: Path, encoding: str) -> str:
    try:
        with path.open("r", encoding=encoding, errors="replace") as stream:
            sample = stream.read(8192)
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except (OSError, csv.Error):
        return ","


def list_sheets(path: Path) -> list[str]:
    if path.suffix.lower() == ".csv":
        return []
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError("UNSUPPORTED_TABLE_FORMAT")
    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    with pd.ExcelFile(path, engine=engine) as book:
        return list(book.sheet_names)


def estimate_table(path: Path, sheet: str | None = None) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError("UNSUPPORTED_TABLE_FORMAT")
    if suffix == ".csv":
        encoding = detect_csv_encoding(path)
        delimiter = sniff_delimiter(path, encoding)
        header = pd.read_csv(path, encoding=encoding, sep=delimiter, nrows=0)
        with path.open("rb") as stream:
            rows = max(sum(1 for _ in stream) - 1, 0)
        columns = len(header.columns)
        sheets: list[str] = []
    else:
        sheets = list_sheets(path)
        selected = sheet or (sheets[0] if len(sheets) == 1 else None)
        if selected is None:
            return {
                "requires_sheet_selection": True,
                "sheets": sheets,
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        if suffix == ".xls":
            import xlrd

            book = xlrd.open_workbook(path, on_demand=True)
            worksheet = book.sheet_by_name(selected)
            rows, columns = max(worksheet.nrows - 1, 0), worksheet.ncols
            book.release_resources()
        else:
            from openpyxl import load_workbook

            book = load_workbook(path, read_only=True, data_only=True)
            worksheet = book[selected]
            rows, columns = max(worksheet.max_row - 1, 0), worksheet.max_column
            book.close()
        sheet = selected
    resource = plan_resources(rows, columns)
    return {
        "requires_sheet_selection": False,
        "rows": int(rows),
        "columns": int(columns),
        "sheet": sheet,
        "sheets": sheets,
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "resource_plan": resource.as_dict(),
    }


def read_table(
    path: Path,
    sheet: str | None = None,
    usecols: Iterable[str] | None = None,
    nrows: int | None = None,
    memory_budget_mb: int = 1536,
) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        encoding = detect_csv_encoding(path)
        estimate = estimate_table(path)
        selected_columns = list(usecols) if usecols is not None else None
        planned_columns = (
            len(selected_columns) if selected_columns is not None else int(estimate["columns"])
        )
        planned_rows = (
            min(int(estimate["rows"]), nrows) if nrows is not None else int(estimate["rows"])
        )
        resource = plan_resources(planned_rows, planned_columns, memory_budget_mb)
        if resource.estimated_memory_bytes > int(resource.budget_bytes * 1.5):
            raise MemoryError("TABLE_REQUIRES_COLUMN_PROJECTION")
        if resource.strategy == "in_memory":
            return pd.read_csv(
                path,
                encoding=encoding,
                sep=sniff_delimiter(path, encoding),
                usecols=selected_columns,
                nrows=nrows,
                low_memory=False,
            )
        chunks = iter_csv_chunks(
            path,
            resource.row_chunk_size,
            usecols=selected_columns,
            nrows=nrows,
        )
        values = list(chunks)
        if not values:
            return pd.read_csv(
                path,
                encoding=encoding,
                sep=sniff_delimiter(path, encoding),
                usecols=selected_columns,
                nrows=0,
            )
        return pd.concat(values, ignore_index=True, copy=False)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        sheets = list_sheets(path)
        selected = sheet or (sheets[0] if len(sheets) == 1 else None)
        if selected is None:
            raise ValueError("EXCEL_SHEET_SELECTION_REQUIRED")
        estimate = estimate_table(path, selected)
        selected_columns = list(usecols) if usecols is not None else None
        planned_columns = (
            len(selected_columns) if selected_columns is not None else int(estimate["columns"])
        )
        planned_rows = (
            min(int(estimate["rows"]), nrows) if nrows is not None else int(estimate["rows"])
        )
        resource = plan_resources(planned_rows, planned_columns, memory_budget_mb)
        if resource.estimated_memory_bytes > int(resource.budget_bytes * 1.5):
            raise MemoryError("TABLE_REQUIRES_COLUMN_PROJECTION")
        return pd.read_excel(
            path,
            sheet_name=selected,
            usecols=selected_columns,
            nrows=nrows,
            engine="xlrd" if suffix == ".xls" else "openpyxl",
        )
    raise ValueError("UNSUPPORTED_TABLE_FORMAT")


def iter_csv_chunks(
    path: Path,
    chunk_size: int,
    *,
    usecols: list[str] | None = None,
    nrows: int | None = None,
) -> Iterable[pd.DataFrame]:
    encoding = detect_csv_encoding(path)
    yield from pd.read_csv(
        path,
        encoding=encoding,
        sep=sniff_delimiter(path, encoding),
        chunksize=chunk_size,
        usecols=usecols,
        nrows=nrows,
        low_memory=False,
    )


def write_table(frame: pd.DataFrame, path: Path, sheet_name: str = "data") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        frame.to_csv(path, index=False, encoding="utf-8-sig")
    elif suffix == ".xlsx":
        frame.to_excel(path, index=False, sheet_name=sheet_name, engine="openpyxl")
    else:
        raise ValueError("UNSUPPORTED_OUTPUT_FORMAT")
    return path


def safe_file_name(name: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in "-_." else "_" for char in name)
    return cleaned.strip("._")[:160] or "data.csv"


def recommended_batches(frame: pd.DataFrame, memory_budget_mb: int = 1536) -> list[list[str]]:
    resource = plan_resources(len(frame), len(frame.columns), memory_budget_mb)
    size = resource.column_batch_size
    columns = list(frame.columns)
    return [columns[index : index + size] for index in range(0, len(columns), size)]


def approximate_file_rows(path: Path) -> int:
    if path.suffix.lower() != ".csv":
        return int(estimate_table(path).get("rows") or 0)
    sample_size = min(path.stat().st_size, 1024 * 1024)
    with path.open("rb") as stream:
        sample = stream.read(sample_size)
    newlines = sample.count(b"\n")
    if not sample_size or not newlines:
        return 0
    return max(0, math.ceil(path.stat().st_size / sample_size * newlines) - 1)
