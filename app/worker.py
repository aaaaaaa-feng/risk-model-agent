from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from sklearn.compose import ColumnTransformer
from sklearn.ensemble import HistGradientBoostingClassifier, RandomForestClassifier
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    average_precision_score,
    brier_score_loss,
    confusion_matrix,
    roc_auc_score,
    roc_curve,
)
from sklearn.base import clone
from sklearn.model_selection import StratifiedKFold, train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler

from .config import MAX_COLUMNS, MAX_ROWS, MAX_UPLOAD_BYTES, MEMORY_BUDGET_BYTES

try:
    from xgboost import XGBClassifier
except Exception:  # pragma: no cover - optional wheel/platform issue
    XGBClassifier = None  # type: ignore[assignment,misc]


LEAKAGE_RE = re.compile(r"(?:post|after|repay|collection|writeoff|settle).*(?:overdue|delinq|default|bad)", re.I)
ID_RE = re.compile(r"(?:^id$|_id$|uuid|phone|mobile|card|identity|身份证|手机号)", re.I)
OOF_MAX_ROWS = 10000


def _parse_datetime(series: pd.Series) -> pd.Series:
    try:
        return pd.to_datetime(series, format="mixed", errors="coerce")
    except (TypeError, ValueError):
        return pd.to_datetime(series, errors="coerce")


def estimate_table_resources(path: Path, sheet: Optional[str] = None) -> Dict[str, Any]:
    """Estimate row/column and in-memory footprint before materializing a table."""
    size_bytes = int(path.stat().st_size)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                if sheet and sheet not in workbook.sheetnames:
                    raise ValueError(f"XLSX_SHEET_NOT_FOUND: {sheet}")
                worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
                rows = int(worksheet.max_row or 0)
                columns = int(worksheet.max_column or 0)
            finally:
                workbook.close()
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError(f"XLSX_RESOURCE_ESTIMATE_FAILED: {exc}") from exc
    elif suffix == ".csv":
        rows = 0
        columns = 0
        decode_errors: List[str] = []
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    sample = handle.read(1024 * 1024)
                    handle.seek(0)
                    if sample:
                        try:
                            import csv

                            dialect = csv.Sniffer().sniff(sample[:10000])
                        except csv.Error:
                            dialect = csv.excel
                        reader = csv.reader(handle, dialect)
                        columns = len(next(reader, []))
                        rows = max(0, sum(1 for _ in reader))
                    break
            except UnicodeDecodeError as exc:
                decode_errors.append(f"{encoding}: {exc}")
        else:
            raise ValueError(f"CSV_ENCODING_UNSUPPORTED: {decode_errors[-1] if decode_errors else 'unknown'}")
    else:
        raise ValueError("UNSUPPORTED_FILE: 仅支持 CSV 和 XLSX")
    data_rows = max(0, rows)
    # A conservative object-heavy estimate; actual pandas memory is measured
    # again after loading. This is a guardrail, not a claim of exact usage.
    estimated_memory_bytes = int(max(size_bytes * 2, data_rows * max(columns, 1) * 96))
    risk = "ok"
    reasons: List[str] = []
    if data_rows > MAX_ROWS:
        risk = "block"
        reasons.append("ROW_LIMIT_EXCEEDED")
    if columns > MAX_COLUMNS:
        risk = "block"
        reasons.append("COLUMN_LIMIT_EXCEEDED")
    if estimated_memory_bytes > MEMORY_BUDGET_BYTES:
        risk = "block"
        reasons.append("MEMORY_BUDGET_EXCEEDED")
    elif estimated_memory_bytes > int(MEMORY_BUDGET_BYTES * 0.7):
        risk = "warn"
        reasons.append("MEMORY_BUDGET_NEAR_LIMIT")
    return {
        "schema_version": "risk-resource-estimate/v1",
        "file_bytes": size_bytes,
        "rows": data_rows,
        "columns": columns,
        "estimated_memory_bytes": estimated_memory_bytes,
        "memory_budget_bytes": MEMORY_BUDGET_BYTES,
        "risk": risk,
        "reasons": reasons,
        "exact": False,
    }


def read_table(path: Path, sheet: Optional[str] = None) -> pd.DataFrame:
    if path.stat().st_size > MAX_UPLOAD_BYTES:
        raise ValueError("FILE_TOO_LARGE: 文件超过本地配置的导入上限")
    resource_estimate = estimate_table_resources(path, sheet)
    if resource_estimate["risk"] == "block":
        reason = ",".join(resource_estimate.get("reasons", [])) or "RESOURCE_LIMIT"
        raise ValueError(f"RESOURCE_LIMIT: 导入前估算超出支持边界（{reason}）")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        frame = pd.read_excel(path, sheet_name=sheet or 0, engine="openpyxl")
    elif suffix == ".csv":
        frame = None
        decode_errors: List[str] = []
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                frame = pd.read_csv(path, nrows=MAX_ROWS + 1, low_memory=True, encoding=encoding)
                break
            except UnicodeDecodeError as exc:
                decode_errors.append(f"{encoding}: {exc}")
        if frame is None:
            raise ValueError(f"CSV_ENCODING_UNSUPPORTED: 无法按 UTF-8/GB18030 解码（{decode_errors[-1] if decode_errors else 'unknown'}）")
    else:
        raise ValueError("UNSUPPORTED_FILE: 仅支持 CSV 和 XLSX")
    if len(frame) > MAX_ROWS:
        raise ValueError("ROW_LIMIT_EXCEEDED: 行数超过当前实测支持上限")
    if len(frame.columns) > MAX_COLUMNS:
        raise ValueError("COLUMN_LIMIT_EXCEEDED: 字段数超过当前设计上限")
    frame.columns = [str(column).strip() or f"unnamed_{index}" for index, column in enumerate(frame.columns)]
    if len(set(frame.columns)) != len(frame.columns):
        raise ValueError("DUPLICATE_COLUMNS: 字段名重复，请先修正表头")
    # Safe, reversible normalization: whitespace-only cells and surrounding
    # whitespace are standardized before profiling. No rows are silently removed.
    object_columns = frame.select_dtypes(include=["object", "string"]).columns
    trimmed_cells = 0
    blank_cells = 0
    for column in object_columns:
        before = frame[column]
        trimmed_cells += int(before.map(lambda value: isinstance(value, str) and value != value.strip()).sum())
        frame[column] = before.map(lambda value: value.strip() if isinstance(value, str) else value)
    if len(object_columns):
        blank_cells = int(frame[object_columns].apply(lambda column: column.astype("string").str.fullmatch(r"\s*")).sum().sum())
        frame[object_columns] = frame[object_columns].replace(r"^\s*$", pd.NA, regex=True)
    frame.attrs["cleaning_audit"] = {
        "trimmed_cells": trimmed_cells,
        "blank_cells_standardized": blank_cells,
        "rows_removed": 0,
        "columns_removed": 0,
    }
    frame.attrs["resource_estimate"] = {**resource_estimate, "exact": True, "actual_memory_bytes": int(frame.memory_usage(deep=True).sum())}
    return frame


def list_sheets(path: Path) -> List[str]:
    if path.suffix.lower() != ".xlsx":
        return []
    if path.stat().st_size > MAX_UPLOAD_BYTES:
        raise ValueError("FILE_TOO_LARGE: 文件超过本地配置的导入上限")
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return [str(name) for name in workbook.sheetnames]
        finally:
            workbook.close()
    except Exception as exc:
        raise ValueError(f"XLSX_SHEET_READ_FAILED: {exc}") from exc


def infer_type(series: pd.Series) -> str:
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_numeric_dtype(series):
        return "numeric"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    parsed = _parse_datetime(series.dropna().head(100))
    if len(parsed) and parsed.notna().mean() > 0.8:
        return "datetime"
    return "categorical" if series.nunique(dropna=True) < min(100, max(10, len(series) // 20)) else "text"


def parse_data_dictionary(frame: pd.DataFrame) -> Dict[str, Any]:
    """Parse a small, local-only field dictionary into a versioned mapping."""
    aliases = {
        "field": {"field", "field_name", "column", "column_name", "name", "字段", "字段名", "变量", "变量名", "特征", "特征名"},
        "display_name": {"display_name", "label", "中文名", "字段中文名", "变量中文名", "展示名"},
        "definition": {"meaning", "description", "definition", "口径", "字段口径", "业务含义", "含义", "定义"},
        "source": {"source", "source_table", "来源", "来源表"},
        "role": {"role", "字段角色", "变量角色", "角色"},
    }
    normalized = {str(column).strip().lower(): column for column in frame.columns}
    selected: Dict[str, Any] = {}
    for key, candidates in aliases.items():
        for candidate, original in normalized.items():
            if candidate in candidates or candidate.replace(" ", "_") in candidates:
                selected[key] = original
                break
    warnings: List[str] = []
    if "field" not in selected:
        if len(frame.columns):
            selected["field"] = frame.columns[0]
            warnings.append("FIELD_COLUMN_INFERRED_FROM_FIRST_COLUMN")
        else:
            return {"schema_version": "risk-data-dictionary/v1", "columns": {}, "field_count": 0, "warnings": ["EMPTY_DICTIONARY"]}
    mapping: Dict[str, Dict[str, Any]] = {}
    for _, row in frame.iterrows():
        raw_name = str(row.get(selected["field"], "") or "").strip()
        if not raw_name or raw_name.lower() == "nan":
            continue
        item: Dict[str, Any] = {}
        for key in ("display_name", "definition", "source", "role"):
            value = row.get(selected.get(key)) if selected.get(key) is not None else None
            if pd.notna(value) and str(value).strip():
                item[key] = str(value).strip()[:500]
        mapping[raw_name] = item
    return {
        "schema_version": "risk-data-dictionary/v1",
        "columns": mapping,
        "field_count": len(mapping),
        "source_columns": [str(column) for column in frame.columns],
        "warnings": warnings,
        "policy": "仅在本机展示原始字段语义；外部 Provider 仍只接收字段别名和聚合证据。",
    }


def profile_table(frame: pd.DataFrame, dictionary: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    columns: List[Dict[str, Any]] = []
    for column in frame.columns:
        series = frame[column]
        unique_count = int(series.nunique(dropna=True))
        missing = float(series.isna().mean())
        candidates = False
        values = set(series.dropna().unique().tolist())
        if values and len(values) <= 2 and values.issubset({0, 1, True, False, "0", "1"}):
            candidates = True
        item = {
            "name": str(column),
            "type": infer_type(series),
            "missing_rate": round(missing, 6),
            "unique_count": unique_count,
            "unique_ratio": round(unique_count / max(len(series), 1), 6),
            "target_candidate": candidates,
            "constant": unique_count <= 1,
        }
        if dictionary and column in (dictionary.get("columns") or {}):
            item["dictionary"] = dictionary["columns"][column]
        columns.append(item)
    return {
        "rows": int(len(frame)),
        "columns": int(len(frame.columns)),
        "memory_bytes": int(frame.memory_usage(deep=True).sum()),
        "resource_estimate": frame.attrs.get("resource_estimate", {}),
        "duplicate_rows": int(frame.duplicated().sum()),
        "target_candidates": [item["name"] for item in columns if item["target_candidate"]],
        "columns_detail": columns,
        "dictionary": {
            "schema_version": (dictionary or {}).get("schema_version"),
            "field_count": (dictionary or {}).get("field_count", 0),
            "matched_count": sum(1 for item in columns if item.get("dictionary")),
            "warnings": (dictionary or {}).get("warnings", []),
        },
        "warnings": _profile_warnings(frame, columns),
        "cleaning": {
            "status": "safe_normalization_only",
            "automatic_actions": ["trim_text_whitespace", "standardize_blank_cells_as_missing"],
            "destructive_actions": [],
            "audit": frame.attrs.get("cleaning_audit", {}),
            "note": "未自动删除样本、截断异常值或改写 Y；训练期填补只在训练分区拟合。",
        },
    }


def quality_analysis(frame: pd.DataFrame, target: Optional[str] = None, time_column: Optional[str] = None) -> Dict[str, Any]:
    """Produce local EDA evidence without emitting customer-level values."""
    numeric: List[Dict[str, Any]] = []
    categorical: List[Dict[str, Any]] = []
    for column in frame.columns:
        series = frame[column]
        missing = int(series.isna().sum())
        if pd.api.types.is_numeric_dtype(series):
            values = series.dropna().astype(float)
            if len(values):
                q1, q3 = values.quantile([0.25, 0.75]).tolist()
                iqr = float(q3 - q1)
                outliers = int(((values < q1 - 1.5 * iqr) | (values > q3 + 1.5 * iqr)).sum()) if iqr else 0
                numeric.append(
                    {
                        "column": str(column),
                        "count": int(len(values)),
                        "missing": missing,
                        "mean": round(float(values.mean()), 6),
                        "median": round(float(values.median()), 6),
                        "min": round(float(values.min()), 6),
                        "max": round(float(values.max()), 6),
                        "p01": round(float(values.quantile(0.01)), 6),
                        "p99": round(float(values.quantile(0.99)), 6),
                        "iqr_outliers": outliers,
                    }
                )
        else:
            categorical.append(
                {
                    "column": str(column),
                    "count": int(series.notna().sum()),
                    "missing": missing,
                    "unique_count": int(series.nunique(dropna=True)),
                    "high_cardinality": bool(series.nunique(dropna=True) / max(len(series), 1) > 0.995),
                }
            )
    time_summary: Dict[str, Any] = {"column": time_column, "valid": False}
    if time_column and time_column in frame.columns:
        parsed = _parse_datetime(frame[time_column])
        if parsed.notna().any():
            time_summary = {
                "column": time_column,
                "valid": bool(parsed.notna().mean() > 0.8),
                "parse_rate": round(float(parsed.notna().mean()), 6),
                "min": str(parsed.min()),
                "max": str(parsed.max()),
            }
    return {
        "schema_version": "risk-eda/v1",
        "rows": int(len(frame)),
        "columns": int(len(frame.columns)),
        "duplicate_rows": int(frame.duplicated().sum()),
        "numeric": numeric,
        "categorical": categorical,
        "target": target_summary(frame, target) if target and target in frame.columns else None,
        "time": time_summary,
        "cleaning_audit": frame.attrs.get("cleaning_audit", {}),
    }


def build_cleaning_plan(frame: pd.DataFrame, profile: Dict[str, Any], quality: Dict[str, Any]) -> Dict[str, Any]:
    actions: List[Dict[str, Any]] = [
        {
            "code": "TRIM_TEXT_AND_BLANKS",
            "category": "safe_auto",
            "status": "applied_on_read",
            "message": "去除文本两端空格，并把空白单元格标准化为缺失。",
        }
    ]
    requires_confirmation: List[Dict[str, Any]] = []
    duplicate_rows = int(quality.get("duplicate_rows", 0))
    if duplicate_rows:
        requires_confirmation.append(
            {
                "code": "DUPLICATE_ROWS_REVIEW",
                "message": f"发现 {duplicate_rows:,} 行完全重复记录；是否去重不能由 Agent 静默决定。",
                "rows": duplicate_rows,
            }
        )
    outlier_columns = [item for item in quality.get("numeric", []) if item.get("iqr_outliers", 0) > 0]
    if outlier_columns:
        requires_confirmation.append(
            {
                "code": "OUTLIER_REVIEW",
                "message": "数值字段存在 IQR 异常值，仅提供证据，不自动截断。",
                "columns": [item["column"] for item in outlier_columns[:100]],
            }
        )
    actions.extend(requires_confirmation)
    return {
        "schema_version": "risk-cleaning-plan/v1",
        "status": "review_required" if requires_confirmation else "safe_normalization_complete",
        "actions": actions,
        "requires_confirmation": requires_confirmation,
        "rows_before": int(len(frame)),
        "rows_after": int(len(frame)),
        "columns_before": int(len(frame.columns)),
        "columns_after": int(len(frame.columns)),
        "rule_version": "cleaning-rules/v1",
        "note": "没有自动删除样本、截断异常值、合并稀有类别或改写 Y。",
    }


def apply_cleaning_plan(frame: pd.DataFrame, plan: Dict[str, Any], approved_actions: Sequence[Any]) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Apply only explicitly approved, parameterized cleaning actions.

    The original frame is never mutated. Every non-trivial action is recorded
    with before/after evidence so a new DatasetVersion can be created by the
    API without replacing the source file.
    """
    requested = list(approved_actions or [])
    allowed = {item.get("code"): item for item in plan.get("requires_confirmation", []) if item.get("code")}
    unknown = []
    normalized: List[Dict[str, Any]] = []
    for item in requested:
        code = item.get("code") if isinstance(item, dict) else item
        if code not in allowed:
            unknown.append(code)
            continue
        normalized.append(item if isinstance(item, dict) else {"code": code})
    if unknown:
        raise ValueError(f"CLEANING_ACTION_NOT_APPROVED: {unknown}")
    result = frame.copy(deep=True)
    evidence: List[Dict[str, Any]] = []
    for action in normalized:
        code = action["code"]
        if code == "DUPLICATE_ROWS_REVIEW":
            before = len(result)
            result = result.drop_duplicates(keep="first").reset_index(drop=True)
            evidence.append({"code": code, "rows_before": before, "rows_after": len(result), "rows_removed": before - len(result)})
        elif code == "OUTLIER_REVIEW":
            columns = action.get("columns") or allowed[code].get("columns") or []
            lower = float(action.get("lower_quantile", 0.01))
            upper = float(action.get("upper_quantile", 0.99))
            if not 0 <= lower < upper <= 1:
                raise ValueError("CLEANING_QUANTILE_INVALID: 分位点必须满足 0 <= lower < upper <= 1")
            clipped: List[Dict[str, Any]] = []
            for column in columns:
                if column not in result.columns or not pd.api.types.is_numeric_dtype(result[column]):
                    raise ValueError(f"CLEANING_COLUMN_INVALID: 异常值处理字段必须是数值列：{column}")
                low, high = result[column].quantile([lower, upper]).tolist()
                changed = int(((result[column] < low) | (result[column] > high)).sum())
                result[column] = result[column].clip(lower=low, upper=high)
                clipped.append({"column": column, "lower": float(low), "upper": float(high), "cells_changed": changed})
            evidence.append({"code": code, "lower_quantile": lower, "upper_quantile": upper, "columns": clipped})
    result.attrs["cleaning_audit"] = {
        **(frame.attrs.get("cleaning_audit") or {}),
        "approved_actions": evidence,
        "rows_removed": int(len(frame) - len(result)),
        "columns_removed": 0,
    }
    return result, {"status": "applied", "actions": evidence, "rows_before": len(frame), "rows_after": len(result), "rows_removed": len(frame) - len(result), "columns_before": len(frame.columns), "columns_after": len(result.columns), "rule_version": plan.get("rule_version", "cleaning-rules/v1")}


def _profile_warnings(frame: pd.DataFrame, columns: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    warnings: List[Dict[str, Any]] = []
    if len(frame) < 200:
        warnings.append({"code": "SMALL_SAMPLE", "severity": "warn", "message": "样本量小于 200，指标波动需要谨慎解释。"})
    if not any(item["target_candidate"] for item in columns):
        warnings.append({"code": "TARGET_NOT_OBVIOUS", "severity": "block", "message": "未发现明确的 0/1 Y 候选字段。"})
    for item in columns:
        if item["missing_rate"] >= 0.95:
            warnings.append({"code": "HIGH_MISSING", "severity": "warn", "column": item["name"], "message": "字段缺失率达到 95% 以上。"})
        if LEAKAGE_RE.search(item["name"]):
            warnings.append({"code": "SUSPECTED_POST_OUTCOME_FEATURE", "severity": "block", "column": item["name"], "message": "字段名疑似包含贷后结果信息，需要业务确认。"})
    return warnings


def target_summary(frame: pd.DataFrame, target: str) -> Dict[str, Any]:
    values = frame[target].value_counts(dropna=False)
    normalized = frame[target].map(_to_binary)
    valid = normalized.notna()
    return {
        "target": target,
        "raw_value_counts": {str(key): int(value) for key, value in values.items()},
        "valid_binary_rows": int(valid.sum()),
        "invalid_rows": int((~valid).sum()),
        "positive_count": int((normalized == 1).sum()),
        "negative_count": int((normalized == 0).sum()),
        "positive_rate": round(float((normalized == 1).mean()), 6),
        "contract_ok": bool(valid.all() and normalized.nunique() == 2),
    }


def _to_binary(value: Any) -> Optional[int]:
    if pd.isna(value):
        return None
    if value in (1, True, "1", "1.0"):
        return 1
    if value in (0, False, "0", "0.0"):
        return 0
    return None


def calculate_iv(frame: pd.DataFrame, target: str, feature: str) -> float:
    y = frame[target].map(_to_binary)
    valid = y.notna()
    x = frame.loc[valid, feature]
    y = y.loc[valid]
    if y.nunique() < 2 or x.nunique(dropna=True) <= 1:
        return 0.0
    if pd.api.types.is_numeric_dtype(x):
        try:
            grouped = pd.qcut(x, q=min(10, max(2, x.nunique())), duplicates="drop")
        except ValueError:
            grouped = x.astype(str)
    else:
        grouped = x.fillna("<MISSING>").astype(str)
        counts = grouped.value_counts()
        small = set(counts[counts < max(10, len(grouped) * 0.01)].index)
        grouped = grouped.map(lambda item: "<OTHER>" if item in small else item)
    table = pd.DataFrame({"group": grouped, "y": y}).groupby("group", observed=False)["y"].agg(["count", "sum"])
    good_total = max(float((y == 0).sum()), 1.0)
    bad_total = max(float((y == 1).sum()), 1.0)
    iv = 0.0
    for row in table.itertuples():
        good_dist = (row.count - row.sum + 0.5) / (good_total + 0.5 * len(table))
        bad_dist = (row.sum + 0.5) / (bad_total + 0.5 * len(table))
        woe = math.log(good_dist / bad_dist)
        iv += (good_dist - bad_dist) * woe
    return float(max(iv, 0.0))


def select_features(
    frame: pd.DataFrame,
    target: str,
    max_features: int = 50,
    min_iv: float = 0.005,
    fit_positions: Optional[Sequence[int]] = None,
    excluded_columns: Optional[Sequence[str]] = None,
) -> Dict[str, Any]:
    """Select variables using only the declared fit rows.

    ``fit_positions`` is deliberately explicit: callers cannot accidentally
    compute IV/WOE on validation or OOT rows while still returning decisions
    for the full schema. This is the main train-only leakage guard in V1.
    """
    fit_frame = frame.iloc[list(fit_positions)] if fit_positions is not None else frame
    excluded = set(excluded_columns or [])
    decisions: List[Dict[str, Any]] = []
    candidates: List[Tuple[str, float]] = []
    for column in frame.columns:
        if column == target:
            continue
        series = fit_frame[column]
        missing = float(series.isna().mean())
        unique_ratio = float(series.nunique(dropna=True) / max(len(series), 1))
        reasons: List[str] = []
        status = "included"
        iv_value: Optional[float] = None
        if column in excluded:
            status, reasons = "excluded", ["USER_EXCLUDED"]
        elif series.nunique(dropna=True) <= 1:
            status, reasons = "excluded", ["CONSTANT"]
        elif missing >= 0.95:
            status, reasons = "excluded", ["HIGH_MISSING"]
        elif ID_RE.search(str(column)) or (
            unique_ratio > 0.995 and not pd.api.types.is_numeric_dtype(series)
        ):
            status, reasons = "excluded", ["SUSPECTED_IDENTIFIER"]
        elif LEAKAGE_RE.search(str(column)):
            status, reasons = "blocked", ["SUSPECTED_POST_OUTCOME_FEATURE"]
        else:
            iv_value = calculate_iv(fit_frame, target, column)
            candidates.append((column, iv_value))
        decisions.append(
            {
                "column": str(column),
                "status": status,
                "missing_rate": round(missing, 6),
                "unique_ratio": round(unique_ratio, 6),
                "iv": round(iv_value, 6) if iv_value is not None else None,
                "reasons": reasons,
            }
        )
    for decision in decisions:
        if decision["status"] == "included" and decision["iv"] is not None and decision["iv"] < min_iv:
            decision["status"] = "excluded"
            decision["reasons"] = ["IV_BELOW_THRESHOLD"]
    candidates = [(column, iv) for column, iv in candidates if iv >= min_iv]
    candidates.sort(key=lambda item: item[1], reverse=True)
    selected = {column for column, _ in candidates[:max_features]}
    for decision in decisions:
        if decision["status"] == "included" and decision["column"] not in selected:
            decision["status"] = "excluded"
            decision["reasons"] = ["IV_RANK_CAPPED"]
    return {
        "selected": [column for column, _ in candidates[:max_features]],
        "decisions": decisions,
        "funnel": {
            "raw": len(frame.columns) - 1,
            "included_after_rules": len(candidates),
            "final": len(selected),
            "blocked": sum(item["status"] == "blocked" for item in decisions),
            "min_iv": min_iv,
            "fit_scope": "train" if fit_positions is not None else "caller_supplied_frame",
            "fit_rows": len(fit_frame),
        },
    }


def split_frame(frame: pd.DataFrame, target: str, time_column: Optional[str] = None) -> Dict[str, Any]:
    y = frame[target].map(_to_binary)
    valid_positions = np.flatnonzero(y.notna().to_numpy())
    work = frame.iloc[valid_positions].copy()
    y = y.iloc[valid_positions].astype(int).to_numpy()
    if time_column and time_column in work.columns:
        parsed = _parse_datetime(work[time_column])
        if parsed.notna().mean() > 0.8:
            order = np.argsort(parsed.fillna(parsed.min()).to_numpy())
            ordered = np.arange(len(work))[order]
            train_end = max(1, int(len(ordered) * 0.6))
            valid_end = max(train_end + 1, int(len(ordered) * 0.8))
            train, valid, oot = ordered[:train_end], ordered[train_end:valid_end], ordered[valid_end:]
            if len(train) < 2 or len(valid) < 2 or len(oot) < 2 or len(np.unique(y[train])) < 2:
                raise ValueError("TIME_SPLIT_CLASS_CONTRACT_FAILED: 时间切分训练集必须包含 0/1 两类且各分区不可为空")
            return {"positions": valid_positions.tolist(), "train": train.tolist(), "valid": valid.tolist(), "oot": oot.tolist(), "method": "time_holdout", "time_column": time_column}
    positions = np.arange(len(work))
    if len(y) < 10 or min(int(np.sum(y == 0)), int(np.sum(y == 1))) < 5:
        raise ValueError("STRATIFIED_SPLIT_CLASS_CONTRACT_FAILED: 分层切分至少需要每类 5 个样本")
    train, remaining = train_test_split(positions, test_size=0.4, random_state=42, stratify=y)
    valid, oot = train_test_split(remaining, test_size=0.5, random_state=42, stratify=y[remaining])
    return {"positions": valid_positions.tolist(), "train": train.tolist(), "valid": valid.tolist(), "oot": oot.tolist(), "method": "stratified_holdout", "time_column": None}


def _preprocessor(frame: pd.DataFrame, features: Sequence[str], dense: bool = False) -> ColumnTransformer:
    numeric = [column for column in features if pd.api.types.is_numeric_dtype(frame[column])]
    categorical = [column for column in features if column not in numeric]
    transformers = []
    if numeric:
        transformers.append(("numeric", Pipeline([("impute", SimpleImputer(strategy="median")), ("scale", StandardScaler())]), numeric))
    if categorical:
        transformers.append(("categorical", Pipeline([("impute", SimpleImputer(strategy="most_frequent")), ("onehot", OneHotEncoder(handle_unknown="ignore", sparse_output=not dense))]), categorical))
    return ColumnTransformer(transformers=transformers, remainder="drop", sparse_threshold=0.2 if not dense else 0.0)


def _metrics(y_true: np.ndarray, probabilities: np.ndarray, threshold: float) -> Dict[str, Any]:
    if len(y_true) == 0 or len(np.unique(y_true)) < 2:
        return {
            "roc_auc": None,
            "pr_auc": None,
            "ks": None,
            "gini": None,
            "brier": None,
            "threshold": threshold,
            "positive_rate": round(float(np.mean(y_true)) if len(y_true) else 0.0, 6),
            "confusion_matrix": None,
            "calibration": [],
        }
    fpr, tpr, thresholds = roc_curve(y_true, probabilities)
    ks = float(np.max(tpr - fpr))
    finite_mask = np.isfinite(thresholds)
    ks_values = np.where(finite_mask, tpr - fpr, -np.inf)
    chosen = float(threshold if threshold is not None else (thresholds[int(np.argmax(ks_values))] if finite_mask.any() else 0.5))
    matrix = confusion_matrix(y_true, (probabilities >= chosen).astype(int), labels=[0, 1]).tolist()
    brier = None
    if np.isfinite(probabilities).all() and np.all((probabilities >= 0) & (probabilities <= 1)):
        brier = round(float(brier_score_loss(y_true, probabilities)), 6)
    return {
        "roc_auc": round(float(roc_auc_score(y_true, probabilities)), 6),
        "pr_auc": round(float(average_precision_score(y_true, probabilities)), 6),
        "ks": round(ks, 6),
        "gini": round(float(2 * roc_auc_score(y_true, probabilities) - 1), 6),
        "brier": brier,
        "threshold": round(chosen, 6),
        "positive_rate": round(float(np.mean(y_true)), 6),
        "confusion_matrix": matrix,
        "calibration": _calibration_table(y_true, probabilities),
    }


def _calibration_table(y_true: np.ndarray, probabilities: np.ndarray, bins: int = 10) -> List[Dict[str, Any]]:
    """Return equal-frequency calibration evidence without selecting a model."""
    if len(y_true) == 0 or len(probabilities) != len(y_true):
        return []
    values = np.asarray(probabilities, dtype=float)
    if not np.isfinite(values).all() or not np.all((values >= 0) & (values <= 1)):
        return []
    order = np.argsort(values, kind="stable")
    chunks = np.array_split(order, min(max(2, bins), len(order)))
    rows: List[Dict[str, Any]] = []
    for index, chunk in enumerate(chunks, start=1):
        if len(chunk) == 0:
            continue
        predicted = float(np.mean(values[chunk]))
        observed = float(np.mean(y_true[chunk]))
        rows.append(
            {
                "bucket": index,
                "row_count": int(len(chunk)),
                "predicted_rate": round(predicted, 6),
                "observed_rate": round(observed, 6),
                "absolute_gap": round(abs(predicted - observed), 6),
            }
        )
    return rows


def _pipeline_feature_importance(pipeline: Pipeline, features: Sequence[str]) -> List[Dict[str, Any]]:
    """Aggregate transformed feature importance back to source columns."""
    preprocess = pipeline.named_steps.get("preprocess")
    model = pipeline.named_steps.get("model")
    if preprocess is None or model is None:
        return []
    try:
        names = [str(item) for item in preprocess.get_feature_names_out()]
    except Exception:
        names = list(features)
    raw_values: Optional[np.ndarray] = None
    if hasattr(model, "feature_importances_"):
        raw_values = np.asarray(getattr(model, "feature_importances_"), dtype=float)
    elif hasattr(model, "coef_"):
        raw_values = np.abs(np.asarray(getattr(model, "coef_"), dtype=float).reshape(-1))
    if raw_values is None or len(raw_values) != len(names):
        return []
    aggregated = {str(feature): 0.0 for feature in features}
    for name, value in zip(names, raw_values):
        source = name.split("__", 1)[-1]
        # One-hot names append the category after the original column name.
        if source not in aggregated:
            source = next((feature for feature in features if source.startswith(f"{feature}_")), source)
        if source in aggregated and np.isfinite(value):
            aggregated[source] += float(abs(value))
    total = sum(aggregated.values())
    rows = [
        {"feature": feature, "importance": round(value, 8), "normalized_importance": round(value / total, 8) if total else 0.0}
        for feature, value in aggregated.items()
    ]
    rows.sort(key=lambda item: (-item["importance"], item["feature"]))
    return rows


def evaluate_baseline(frame: pd.DataFrame, target: str, score_column: str, split: Dict[str, Any]) -> Dict[str, Any]:
    """Evaluate an existing probability/score column on the frozen split."""
    if score_column not in frame.columns:
        raise ValueError(f"BASELINE_COLUMN_NOT_FOUND: {score_column}")
    values = pd.to_numeric(frame.iloc[split["positions"]][score_column], errors="coerce").to_numpy(dtype=float)
    y = frame.iloc[split["positions"]][target].map(_to_binary).astype(int).to_numpy()
    if not np.isfinite(values).all():
        raise ValueError("BASELINE_SCORE_INVALID: 基线分数必须全部是可解析数值")
    train_idx = np.asarray(split["train"], dtype=int)
    valid_idx = np.asarray(split["valid"], dtype=int)
    oot_idx = np.asarray(split["oot"], dtype=int)
    # A baseline may be a score where larger means safer. Choose orientation
    # once on the validation portion and freeze it before OOT evaluation.
    direct = _metrics(y[valid_idx], values[valid_idx], None)
    inverse = _metrics(y[valid_idx], -values[valid_idx], None)
    orientation = "higher_is_bad" if (direct.get("roc_auc") or -1) >= (inverse.get("roc_auc") or -1) else "higher_is_good"
    probabilities = values if orientation == "higher_is_bad" else -values
    threshold = _metrics(y[valid_idx], probabilities[valid_idx], None)["threshold"]
    approval_rate = 0.8
    return {
        "name": "baseline",
        "score_column": score_column,
        "orientation": orientation,
        "validation": _metrics(y[valid_idx], probabilities[valid_idx], threshold),
        "train": _metrics(y[train_idx], probabilities[train_idx], threshold),
        "oot": _metrics(y[oot_idx], probabilities[oot_idx], threshold),
        "validation_lift": _lift_table(y[valid_idx], probabilities[valid_idx]),
        "oot_lift": _lift_table(y[oot_idx], probabilities[oot_idx]),
        "validation_fixed_rate": _fixed_rate_metrics(y[valid_idx], probabilities[valid_idx], approval_rate),
        "oot_fixed_rate": _fixed_rate_metrics(y[oot_idx], probabilities[oot_idx], approval_rate),
        "fixed_approval_rate": approval_rate,
        "protocol": "risk-validation/holdout-v1",
        "oot_used_for_selection": False,
    }


def reevaluate_baseline(
    frame: pd.DataFrame,
    target: str,
    score_column: str,
    orientation: str,
    threshold: float,
    approval_rate: float = 0.8,
) -> Dict[str, Any]:
    """Evaluate an existing score column on a new OOT-only dataset.

    Direction and threshold must come from the already-frozen baseline
    validation result. The new dataset contributes evaluation evidence only;
    it cannot re-orient the score, choose a threshold, or affect the champion.
    """
    if score_column not in frame.columns:
        raise ValueError(f"BASELINE_COLUMN_NOT_FOUND: {score_column}")
    if target not in frame.columns:
        raise ValueError(f"TARGET_COLUMN_NOT_FOUND: {target}")
    if orientation not in {"higher_is_bad", "higher_is_good"}:
        raise ValueError("BASELINE_ORIENTATION_INVALID")
    try:
        frozen_threshold = float(threshold)
        rate = float(approval_rate)
    except (TypeError, ValueError) as exc:
        raise ValueError("BASELINE_REEVALUATION_PARAMETER_INVALID") from exc
    if not np.isfinite(frozen_threshold) or not 0 < rate <= 1:
        raise ValueError("BASELINE_REEVALUATION_PARAMETER_INVALID")
    normalized = frame[target].map(_to_binary)
    if normalized.isna().any() or len(normalized) < 2 or normalized.nunique() < 2:
        raise ValueError("BASELINE_REEVALUATION_TARGET_INVALID: 新 OOT 必须包含完整的 0/1 两类 Y")
    values = pd.to_numeric(frame[score_column], errors="coerce").to_numpy(dtype=float)
    if not np.isfinite(values).all():
        raise ValueError("BASELINE_SCORE_INVALID: 基线分数必须全部是可解析数值")
    y = normalized.astype(int).to_numpy()
    risk_scores = values if orientation == "higher_is_bad" else -values
    return {
        "schema_version": "risk-baseline-reevaluation/v1",
        "rows": int(len(frame)),
        "target": target,
        "score_column": score_column,
        "orientation": orientation,
        "threshold": round(frozen_threshold, 8),
        "approval_rate": round(rate, 6),
        "metrics": _metrics(y, risk_scores, frozen_threshold),
        "lift": _lift_table(y, risk_scores),
        "fixed_rate": _fixed_rate_metrics(y, risk_scores, rate),
        "fit_scope": "none",
        "eval_scope": "new_oot_only",
        "oot_used_for_selection": False,
        "fact_boundary": "仅为新 OOT 离线复评，不改变正式冠军或阈值",
    }


def _lift_table(y_true: np.ndarray, probabilities: np.ndarray, bins: int = 10) -> List[Dict[str, Any]]:
    if len(y_true) == 0:
        return []
    order = np.argsort(-probabilities, kind="stable")
    chunks = np.array_split(order, min(bins, len(order)))
    total_bad = int(np.sum(y_true))
    base_rate = float(np.mean(y_true)) if len(y_true) else 0.0
    cumulative_bad = 0
    rows: List[Dict[str, Any]] = []
    for index, chunk in enumerate(chunks, start=1):
        if len(chunk) == 0:
            continue
        bad_count = int(np.sum(y_true[chunk]))
        cumulative_bad += bad_count
        response_rate = bad_count / len(chunk)
        rows.append(
            {
                "bucket": index,
                "row_count": int(len(chunk)),
                "bad_count": bad_count,
                "response_rate": round(response_rate, 6),
                "lift": round(response_rate / base_rate, 6) if base_rate else None,
                "cumulative_capture": round(cumulative_bad / total_bad, 6) if total_bad else None,
            }
        )
    return rows


def _fixed_rate_metrics(y_true: np.ndarray, risk_scores: np.ndarray, approval_rate: float = 0.8) -> Dict[str, Any]:
    """Aggregate approval/rejection evidence at a frozen approval rate."""
    if len(y_true) == 0:
        return {"approval_rate": approval_rate, "approved_count": 0, "rejected_count": 0, "bad_capture_rate": None, "approved_bad_rate": None}
    count = min(len(y_true), max(0, int(round(len(y_true) * approval_rate))))
    order = np.argsort(np.asarray(risk_scores, dtype=float))
    approved = np.zeros(len(y_true), dtype=bool)
    approved[order[:count]] = True
    rejected = ~approved
    bad_total = int(np.sum(y_true))
    return {
        "approval_rate": round(float(count / len(y_true)), 6),
        "approved_count": int(np.sum(approved)),
        "rejected_count": int(np.sum(rejected)),
        "approved_bad_rate": round(float(np.mean(y_true[approved])) if approved.any() else 0.0, 6),
        "rejected_bad_rate": round(float(np.mean(y_true[rejected])) if rejected.any() else 0.0, 6),
        "bad_capture_rate": round(float(np.sum(y_true[rejected]) / bad_total), 6) if bad_total else None,
    }


def _swap_set(y_true: np.ndarray, baseline_scores: np.ndarray, candidate_scores: np.ndarray, approval_rate: float = 0.8) -> Dict[str, Any]:
    def approved(scores: np.ndarray) -> np.ndarray:
        count = min(len(scores), max(0, int(round(len(scores) * approval_rate))))
        mask = np.zeros(len(scores), dtype=bool)
        mask[np.argsort(scores)[:count]] = True
        return mask

    baseline_approved = approved(baseline_scores)
    candidate_approved = approved(candidate_scores)
    groups = {
        "both_approved": baseline_approved & candidate_approved,
        "both_rejected": ~baseline_approved & ~candidate_approved,
        "baseline_approved_candidate_rejected": baseline_approved & ~candidate_approved,
        "baseline_rejected_candidate_approved": ~baseline_approved & candidate_approved,
    }
    result: Dict[str, Any] = {"approval_rate": approval_rate, "groups": {}}
    for name, mask in groups.items():
        count = int(np.sum(mask))
        result["groups"][name] = {"count": count, "bad_count": int(np.sum(y_true[mask])) if count else 0, "bad_rate": round(float(np.mean(y_true[mask])), 6) if count else None}
    return result


def _psi_from_counts(reference: np.ndarray, current: np.ndarray) -> float:
    reference = np.asarray(reference, dtype=float)
    current = np.asarray(current, dtype=float)
    reference = reference / max(float(reference.sum()), 1.0)
    current = current / max(float(current.sum()), 1.0)
    reference = np.clip(reference, 1e-6, None)
    current = np.clip(current, 1e-6, None)
    return float(np.sum((current - reference) * np.log(current / reference)))


def _stability_for_feature(reference: pd.Series, current: pd.Series) -> Dict[str, Any]:
    """Compute train-fitted PSI bins for one feature and one evaluation split."""
    if pd.api.types.is_numeric_dtype(reference):
        ref = pd.to_numeric(reference, errors="coerce")
        cur = pd.to_numeric(current, errors="coerce")
        finite = ref.dropna().astype(float)
        if finite.nunique() > 1:
            edges = np.unique(np.nanquantile(finite.to_numpy(), np.linspace(0, 1, 11))).astype(float)
            if len(edges) > 1:
                edges[0] = -np.inf
                edges[-1] = np.inf
            else:
                edges = np.array([-np.inf, np.inf])
        else:
            edges = np.array([-np.inf, np.inf])
        ref_labels = pd.cut(ref, bins=edges, include_lowest=True, duplicates="drop").astype("string").fillna("<MISSING>")
        cur_labels = pd.cut(cur, bins=edges, include_lowest=True, duplicates="drop").astype("string").fillna("<MISSING>")
        categories = sorted(set(ref_labels.tolist()) | set(cur_labels.tolist()))
    else:
        ref_labels = reference.fillna("<MISSING>").astype(str)
        cur_labels = current.fillna("<MISSING>").astype(str)
        top = list(ref_labels.value_counts().head(20).index)
        categories = list(dict.fromkeys(top + ["<OTHER>"]))
        ref_labels = ref_labels.map(lambda value: value if value in top else "<OTHER>")
        cur_labels = cur_labels.map(lambda value: value if value in top else "<OTHER>")
    ref_counts = ref_labels.value_counts().reindex(categories, fill_value=0).to_numpy(dtype=float)
    cur_counts = cur_labels.value_counts().reindex(categories, fill_value=0).to_numpy(dtype=float)
    return {
        "psi": round(_psi_from_counts(ref_counts, cur_counts), 6),
        "reference_rows": int(len(reference)),
        "current_rows": int(len(current)),
        "bins": int(len(categories)),
        "fit_scope": "train",
        "note": "分箱/类别集合仅使用训练分区拟合；PSI 阈值只是复核提示，不是普遍定律。",
    }


def stability_analysis(frame: pd.DataFrame, features: Sequence[str], split: Dict[str, Any]) -> Dict[str, Any]:
    """Summarize train-only PSI and correlation evidence for selected features."""
    train_idx = np.asarray(split.get("train", []), dtype=int)
    valid_idx = np.asarray(split.get("valid", []), dtype=int)
    oot_idx = np.asarray(split.get("oot", []), dtype=int)
    result: Dict[str, Any] = {"schema_version": "risk-stability/v1", "fit_scope": "train", "features": [], "correlation": []}
    for feature in features:
        if feature not in frame.columns:
            continue
        reference = frame.iloc[train_idx][feature]
        valid = frame.iloc[valid_idx][feature]
        oot = frame.iloc[oot_idx][feature]
        item = {
            "feature": str(feature),
            "validation": _stability_for_feature(reference, valid),
            "oot": _stability_for_feature(reference, oot),
        }
        for scope in ("validation", "oot"):
            value = item[scope]["psi"]
            item[scope]["review_flag"] = "high" if value >= 0.25 else ("review" if value >= 0.1 else "ok")
        result["features"].append(item)
    numeric = [feature for feature in features if feature in frame.columns and pd.api.types.is_numeric_dtype(frame[feature])]
    if len(numeric) >= 2 and len(train_idx):
        corr = frame.iloc[train_idx][numeric].corr().abs()
        for left_index, left in enumerate(numeric):
            for right in numeric[left_index + 1 :]:
                value = corr.loc[left, right]
                if pd.notna(value) and float(value) >= 0.7:
                    result["correlation"].append({"feature_a": left, "feature_b": right, "absolute_correlation": round(float(value), 6), "fit_scope": "train", "review_flag": "high" if value >= 0.9 else "review"})
        result["correlation"].sort(key=lambda item: (-item["absolute_correlation"], item["feature_a"], item["feature_b"]))
        result["correlation"] = result["correlation"][:100]
    return result


def _oof_diagnostic(
    pipeline: Pipeline,
    X: pd.DataFrame,
    y: np.ndarray,
    train_idx: np.ndarray,
    folds: int = 3,
) -> Dict[str, Any]:
    """Fit preprocessing independently inside train-only folds.

    This is diagnostic OOF evidence, never a replacement for the frozen
    validation set used to choose the champion. Large datasets skip it with an
    explicit reason so an 8GB machine does not silently overcommit memory.
    """
    y_train = y[train_idx]
    if len(train_idx) > OOF_MAX_ROWS:
        return {"status": "skipped", "reason": "OOF_ROW_CAP", "max_rows": OOF_MAX_ROWS}
    if min(int(np.sum(y_train == 0)), int(np.sum(y_train == 1))) < folds:
        return {"status": "skipped", "reason": "OOF_CLASS_COUNT_TOO_SMALL", "folds": folds}
    splitter = StratifiedKFold(n_splits=folds, shuffle=True, random_state=42)
    probabilities = np.full(len(train_idx), np.nan, dtype=float)
    for fit_rel, eval_rel in splitter.split(train_idx, y_train):
        fold = clone(pipeline)
        fold.fit(X.iloc[train_idx[fit_rel]], y_train[fit_rel])
        probabilities[eval_rel] = fold.predict_proba(X.iloc[train_idx[eval_rel]])[:, 1]
    metrics = _metrics(y_train, probabilities, None)
    return {"status": "succeeded", "folds": folds, "coverage": round(float(np.isfinite(probabilities).mean()), 6), "metrics": metrics}


def _woe_oof_diagnostic(
    frame: pd.DataFrame,
    target: str,
    features: Sequence[str],
    y: np.ndarray,
    train_idx: np.ndarray,
    folds: int = 3,
) -> Dict[str, Any]:
    """OOF diagnostic with WOE bins refit inside every train-only fold."""
    y_train = y[train_idx]
    if len(train_idx) > OOF_MAX_ROWS:
        return {"status": "skipped", "reason": "OOF_ROW_CAP", "max_rows": OOF_MAX_ROWS}
    if min(int(np.sum(y_train == 0)), int(np.sum(y_train == 1))) < folds:
        return {"status": "skipped", "reason": "OOF_CLASS_COUNT_TOO_SMALL", "folds": folds}
    splitter = StratifiedKFold(n_splits=folds, shuffle=True, random_state=42)
    probabilities = np.full(len(train_idx), np.nan, dtype=float)
    for fit_rel, eval_rel in splitter.split(train_idx, y_train):
        fit_idx = train_idx[fit_rel]
        eval_idx = train_idx[eval_rel]
        specs = _fit_woe_specs(frame, target, features, fit_idx)
        transformed = pd.DataFrame(
            {
                feature: _woe_labels(frame[feature], spec).map(spec["woe"]).fillna(0.0)
                for feature, spec in specs.items()
            }
        )
        model = LogisticRegression(max_iter=500, class_weight="balanced", random_state=42)
        model.fit(transformed.iloc[fit_idx], y[fit_idx])
        probabilities[eval_rel] = model.predict_proba(transformed.iloc[eval_idx])[:, 1]
    metrics = _metrics(y_train, probabilities, None)
    return {"status": "succeeded", "folds": folds, "coverage": round(float(np.isfinite(probabilities).mean()), 6), "metrics": metrics}


def train_candidates(
    frame: pd.DataFrame,
    target: str,
    features: Sequence[str],
    split: Dict[str, Any],
    output_dir: Path,
    model_names: Optional[Sequence[str]] = None,
    baseline_column: Optional[str] = None,
) -> Dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    valid_frame = frame.iloc[split["positions"]].reset_index(drop=True)
    y = valid_frame[target].map(_to_binary).astype(int).to_numpy()
    X = valid_frame[list(features)].copy()
    train_idx = np.asarray(split["train"], dtype=int)
    valid_idx = np.asarray(split["valid"], dtype=int)
    oot_idx = np.asarray(split["oot"], dtype=int)
    train_positive = int(np.sum(y[train_idx] == 1))
    train_negative = int(np.sum(y[train_idx] == 0))
    # split_frame enforces both classes for supported runs; keep the fallback
    # deterministic for direct Worker callers while reporting true counts.
    scale_pos_weight = round(train_negative / train_positive, 8) if train_positive else 1.0
    imbalance_policy = {
        "schema_version": "risk-imbalance-policy/v1",
        "fit_scope": "train",
        "train_positive_count": train_positive,
        "train_negative_count": train_negative,
        "policy": "algorithmic_class_weight",
        "resampling": "none",
    }
    all_models: List[Tuple[str, Any, bool]] = [
        ("logistic_regression", LogisticRegression(max_iter=500, class_weight="balanced", random_state=42), False),
        ("random_forest", RandomForestClassifier(n_estimators=120, max_depth=8, min_samples_leaf=3, class_weight="balanced_subsample", n_jobs=1, random_state=42), True),
        ("hist_gradient_boosting", HistGradientBoostingClassifier(max_iter=120, learning_rate=0.06, max_leaf_nodes=15, class_weight="balanced", random_state=42), True),
    ]
    all_models.append(
        (
            "xgboost",
            XGBClassifier(n_estimators=140, max_depth=4, learning_rate=0.07, subsample=0.85, colsample_bytree=0.85, scale_pos_weight=scale_pos_weight, n_jobs=1, eval_metric="logloss", random_state=42) if XGBClassifier is not None else None,
            True,
        )
    )
    requested = list(dict.fromkeys(model_names or [name for name, _, _ in all_models] + ["woe_logistic_scorecard"]))
    allowed = {name for name, _, _ in all_models} | {"woe_logistic_scorecard"}
    unknown = [name for name in requested if name not in allowed]
    if unknown:
        raise ValueError(f"UNKNOWN_MODEL: 不支持的候选模型 {unknown}")
    models = [(name, estimator, dense) for name, estimator, dense in all_models if name in requested]
    scorecard_enabled = "woe_logistic_scorecard" in requested
    candidates: List[Dict[str, Any]] = []
    validation_probabilities: Dict[str, np.ndarray] = {}
    scorecard = _scorecard_from_woe(valid_frame, target, features, train_idx, valid_idx, oot_idx)
    if scorecard_enabled and scorecard.get("route") == "woe_logistic":
        validation_probabilities["woe_logistic_scorecard"] = np.asarray(scorecard.get("_validation_probability", []), dtype=float)
        candidates.append(
            {
                "name": "woe_logistic_scorecard",
                "status": "succeeded",
                "features": len(features),
                "validation_protocol": {
                    "version": "risk-validation/holdout-v1",
                    "fit_scope": "train",
                    "tuning_scope": "none",
                    "validation_eval_scope": "validation",
                    "oot_eval_scope": "oot",
                    "oot_used_for_selection": False,
                },
                "validation": scorecard["validation"],
                "train": scorecard["train"],
                "oot": scorecard["oot"],
                "validation_lift": scorecard["validation_lift"],
                "oot_lift": scorecard["oot_lift"],
                "validation_fixed_rate": _fixed_rate_metrics(y[valid_idx], validation_probabilities["woe_logistic_scorecard"]),
                "oot_fixed_rate": _fixed_rate_metrics(y[oot_idx], np.asarray(scorecard.get("_oot_probability", []), dtype=float)),
                "oof": scorecard.get("oof"),
                "params": scorecard.get("params", {}),
                "feature_importance": scorecard.get("feature_importance", []),
                "score_mapping_check": scorecard.get("score_mapping_check", {}),
            }
        )
        scorecard.pop("_validation_probability", None)
        scorecard.pop("_oot_probability", None)
    for name, estimator, dense in models:
        try:
            if estimator is None:
                raise RuntimeError("XGBOOST_UNAVAILABLE: 当前 Python 环境未安装可用的 XGBoost")
            pipeline = Pipeline([("preprocess", _preprocessor(X, features, dense=dense)), ("model", estimator)])
            oof = _oof_diagnostic(pipeline, X, y, train_idx)
            pipeline.fit(X.iloc[train_idx], y[train_idx])
            valid_probability = pipeline.predict_proba(X.iloc[valid_idx])[:, 1]
            validation_probabilities[name] = valid_probability
            threshold_metrics = _metrics(y[valid_idx], valid_probability, None)
            threshold = threshold_metrics["threshold"] if threshold_metrics["threshold"] is not None else 0.5
            train_probability = pipeline.predict_proba(X.iloc[train_idx])[:, 1]
            oot_probability = pipeline.predict_proba(X.iloc[oot_idx])[:, 1] if len(oot_idx) else np.array([])
            item = {
                "name": name,
                "status": "succeeded",
                "features": len(features),
                "validation_protocol": {
                    "version": "risk-validation/holdout-v1",
                    "fit_scope": "train",
                    "tuning_scope": "none",
                    "validation_eval_scope": "validation",
                    "oot_eval_scope": "oot",
                    "oot_used_for_selection": False,
                },
                "validation": _metrics(y[valid_idx], valid_probability, threshold),
                "train": _metrics(y[train_idx], train_probability, threshold),
                "oot": _metrics(y[oot_idx], oot_probability, threshold),
                "validation_lift": _lift_table(y[valid_idx], valid_probability),
                "oot_lift": _lift_table(y[oot_idx], oot_probability),
                "validation_fixed_rate": _fixed_rate_metrics(y[valid_idx], valid_probability),
                "oot_fixed_rate": _fixed_rate_metrics(y[oot_idx], oot_probability),
                "oof": oof,
                "params": estimator.get_params(deep=False),
                "feature_importance": _pipeline_feature_importance(pipeline, features),
            }
            candidates.append(item)
            try:
                if name == "xgboost" and hasattr(estimator, "save_model"):
                    native_path = output_dir / f"{name}.json"
                    estimator.save_model(native_path)
                    item["serialization"] = {
                        "format": "xgboost_native_json",
                        "artifact": native_path.name,
                        "load_policy": "仅从本 Run 本地目录加载；不接受外部不受信模型文件。",
                    }
                else:
                    import joblib

                    artifact = output_dir / f"{name}.joblib"
                    joblib.dump(pipeline, artifact)
                    item["serialization"] = {
                        "format": "joblib_local_pipeline",
                        "artifact": artifact.name,
                        "load_policy": "仅从本 Run 本地目录加载；不接受外部不受信模型文件。",
                    }
            except Exception:
                item["serialization"] = {"format": "unavailable", "artifact": None, "load_policy": "模型文件未成功写入，不能宣称可加载。"}
        except Exception as exc:  # candidate isolation is intentional
            candidates.append({"name": name, "status": "failed", "error": f"{type(exc).__name__}: {exc}", "features": len(features)})
    succeeded = [item for item in candidates if item["status"] == "succeeded"]
    champion = max(succeeded, key=lambda item: (item["validation"]["roc_auc"] or -1, item["validation"]["ks"] or -1), default=None)
    stability = stability_analysis(valid_frame, features, {"train": train_idx, "valid": valid_idx, "oot": oot_idx})
    baseline = None
    if baseline_column:
        baseline = evaluate_baseline(valid_frame, target, baseline_column, split)
        if champion and champion.get("name") in validation_probabilities:
            baseline_scores = pd.to_numeric(valid_frame[baseline_column], errors="coerce").to_numpy(dtype=float)
            if baseline.get("orientation") == "higher_is_good":
                baseline_scores = -baseline_scores
            baseline["swap_set"] = _swap_set(y[valid_idx], baseline_scores[valid_idx], validation_probabilities[champion["name"]], baseline.get("fixed_approval_rate", 0.8))
    return {
        "split": {key: value for key, value in split.items() if key != "positions"},
        "models_requested": requested,
        "candidates": candidates,
        "champion": champion,
        "scorecard": scorecard,
        "baseline": baseline,
        "stability": stability,
        "imbalance_policy": imbalance_policy,
    }


def _woe_labels(series: pd.Series, spec: Dict[str, Any]) -> pd.Series:
    if spec["kind"] == "numeric":
        labels = pd.cut(series.astype(float), bins=spec["edges"], include_lowest=True, duplicates="drop").astype("string")
        return labels.fillna("<MISSING>")
    values = series.fillna("<MISSING>").astype(str)
    categories = set(spec.get("categories", []))
    return values.map(lambda value: value if value in categories else "<OTHER>")


def _fit_woe_specs(frame: pd.DataFrame, target: str, features: Sequence[str], train_idx: np.ndarray) -> Dict[str, Any]:
    train = frame.iloc[train_idx]
    y = train[target].map(_to_binary).astype(int)
    specs: Dict[str, Any] = {}
    for feature in features:
        series = train[feature]
        if pd.api.types.is_numeric_dtype(series):
            finite = series.dropna().astype(float)
            if finite.nunique() > 1:
                quantiles = np.nanquantile(finite.to_numpy(), np.linspace(0, 1, min(11, finite.nunique() + 1)))
                edges = np.unique(quantiles).astype(float).tolist()
                if len(edges) < 2:
                    edges = [float(finite.min()), float(finite.max())]
                edges[0] = -np.inf
                edges[-1] = np.inf
            else:
                edges = [-np.inf, np.inf]
            kind = "numeric"
            categories: List[str] = []
        else:
            values = series.fillna("<MISSING>").astype(str)
            counts = values.value_counts()
            categories = [str(item) for item in counts.head(50).index]
            if "<MISSING>" not in categories and values.eq("<MISSING>").any():
                categories.append("<MISSING>")
            edges = []
            kind = "categorical"
        spec = {"kind": kind, "edges": edges, "categories": categories}
        labels = _woe_labels(series, spec)
        table = pd.DataFrame({"group": labels, "y": y.to_numpy()}).groupby("group", observed=False)["y"].agg(["count", "sum"])
        good_total = max(float((y == 0).sum()), 1.0)
        bad_total = max(float((y == 1).sum()), 1.0)
        mapping: Dict[str, float] = {}
        rows: List[Dict[str, Any]] = []
        for row in table.itertuples():
            good = float(row.count - row.sum)
            bad = float(row.sum)
            good_dist = (good + 0.5) / (good_total + 0.5 * len(table))
            bad_dist = (bad + 0.5) / (bad_total + 0.5 * len(table))
            woe = float(math.log(good_dist / bad_dist))
            iv = float((good_dist - bad_dist) * woe)
            label = str(row.Index)
            mapping[label] = woe
            rows.append({"bin": label, "count": int(row.count), "good_count": int(good), "bad_count": int(bad), "bad_rate": round(bad / max(float(row.count), 1.0), 6), "woe": round(woe, 6), "iv": round(iv, 6)})
        spec["woe"] = mapping
        spec["iv"] = round(sum(item["iv"] for item in rows), 6)
        spec["rows"] = rows
        specs[feature] = spec
    return specs


def _scorecard_from_woe(
    frame: pd.DataFrame,
    target: str,
    features: Sequence[str],
    train_idx: np.ndarray,
    valid_idx: np.ndarray,
    oot_idx: np.ndarray,
) -> Dict[str, Any]:
    try:
        y = frame[target].map(_to_binary).astype(int).to_numpy()
        specs = _fit_woe_specs(frame, target, features, train_idx)
        transformed = pd.DataFrame({feature: _woe_labels(frame[feature], spec).map(spec["woe"]).fillna(0.0) for feature, spec in specs.items()})
        model = LogisticRegression(max_iter=500, class_weight="balanced", random_state=42)
        model.fit(transformed.iloc[train_idx], y[train_idx])
        train_probability = model.predict_proba(transformed.iloc[train_idx])[:, 1]
        valid_probability = model.predict_proba(transformed.iloc[valid_idx])[:, 1]
        oot_probability = model.predict_proba(transformed.iloc[oot_idx])[:, 1] if len(oot_idx) else np.array([])
        threshold = _metrics(y[valid_idx], valid_probability, None)["threshold"]
        base_score = 600.0
        pdo = 20.0
        odds = 50.0
        factor = pdo / math.log(2)
        intercept = float(model.intercept_[0])
        coefficients = {feature: float(coef) for feature, coef in zip(features, model.coef_[0])}
        base_points = base_score + factor * (-intercept - math.log(odds))
        points: List[Dict[str, Any]] = []
        for feature in features:
            coefficient = coefficients[feature]
            for row in specs[feature]["rows"]:
                points.append({"feature": feature, **row, "coefficient": round(coefficient, 6), "points": round(-factor * coefficient * specs[feature]["woe"][row["bin"]], 4)})
        score_values = np.full(len(transformed), base_points, dtype=float)
        for feature in features:
            score_values += -factor * coefficients[feature] * transformed[feature].to_numpy(dtype=float)
        score_odds = odds * np.power(2.0, (score_values - base_score) / pdo)
        score_probability = 1.0 / (1.0 + score_odds)
        model_probability = model.predict_proba(transformed)[:, 1]
        mapping_error = np.abs(score_probability - model_probability)
        return {
            "route": "woe_logistic",
            "base_score": base_score,
            "pdo": pdo,
            "odds": odds,
            "factor": round(factor, 6),
            "base_points": round(base_points, 4),
            "formula": "score = base_points + sum(bin_points); odds_good_to_bad = odds * 2 ** ((score - base_score) / pdo)",
            "features": list(features),
            "bins": {feature: {key: value for key, value in spec.items() if key not in {"woe", "rows"}} | {"iv": spec["iv"], "rows": spec["rows"]} for feature, spec in specs.items()},
            "points": points,
            "feature_importance": [
                {"feature": feature, "coefficient": round(coefficients[feature], 8), "absolute_coefficient": round(abs(coefficients[feature]), 8), "iv": specs[feature]["iv"]}
                for feature in sorted(features, key=lambda item: (-abs(coefficients[item]), item))
            ],
            "score_mapping_check": {
                "passed": bool(np.isfinite(mapping_error).all() and (float(np.max(mapping_error)) if len(mapping_error) else 0.0) <= 1e-6),
                "sample_count": int(len(score_values)),
                "max_absolute_probability_error": round(float(np.max(mapping_error)) if len(mapping_error) else 0.0, 10),
                "score_min": round(float(np.min(score_values)) if len(score_values) else 0.0, 6),
                "score_max": round(float(np.max(score_values)) if len(score_values) else 0.0, 6),
                "checked_scopes": ["train", "validation", "oot"],
            },
            "params": {"max_iter": 500, "class_weight": "balanced", "random_state": 42},
            "validation": _metrics(y[valid_idx], valid_probability, threshold),
            "train": _metrics(y[train_idx], train_probability, threshold),
            "oot": _metrics(y[oot_idx], oot_probability, threshold),
            "validation_lift": _lift_table(y[valid_idx], valid_probability),
            "oot_lift": _lift_table(y[oot_idx], oot_probability),
            "_validation_probability": valid_probability,
            "_oot_probability": oot_probability,
            "oof": _woe_oof_diagnostic(frame, target, features, y, train_idx),
            "fit_scope": "train",
            "note": "WOE 分箱和 Logistic 系数只在训练分区拟合；验证/OOT 仅用于评估。",
        }
    except Exception as exc:
        return {"route": "unavailable", "error": f"{type(exc).__name__}: {exc}"}


def segment_analysis(frame: pd.DataFrame, spec: Dict[str, Any]) -> Dict[str, Any]:
    dimensions = spec.get("dimensions") or []
    if not 1 <= len(dimensions) <= 4:
        raise ValueError("INVALID_ANALYSIS_SPEC: 维度必须为 1—4 个")
    max_groups = max(1, min(int(spec.get("max_groups", 1000)), 5000))
    top_k = max(2, min(int(spec.get("top_k_per_dimension", 20)), 100))
    columns = [item.get("column") for item in dimensions]
    missing = [column for column in columns if column not in frame.columns]
    if missing:
        raise ValueError(f"INVALID_ANALYSIS_SPEC: 字段不存在 {missing}")
    target = spec.get("target", {}).get("column")
    work = frame.copy()
    if target in work.columns:
        work["__target__"] = work[target].map(_to_binary)
    for item in dimensions:
        column = item["column"]
        if item.get("transform") in ("quantile_bins", "numeric_bins") and pd.api.types.is_numeric_dtype(work[column]):
            work[f"__dim_{column}"] = pd.qcut(work[column], q=min(int(item.get("bins", 10)), 20), duplicates="drop")
        else:
            values = work[column].fillna("<MISSING>").astype(str)
            counts = values.value_counts(dropna=False)
            keep = set(counts.head(top_k).index)
            work[f"__dim_{column}"] = values.map(lambda value: value if value in keep else "<OTHER>")
    group_cols = [f"__dim_{column}" for column in columns]
    estimated_groups = 1
    for group_column in group_cols:
        estimated_groups *= max(1, int(work[group_column].nunique(dropna=False)))
        if estimated_groups > max_groups * 20:
            raise ValueError(f"GROUP_LIMIT_EXCEEDED: 预计组合数 {estimated_groups:,} 超过当前上限，请减少维度或提高筛选条件")
    grouped = work.groupby(group_cols, observed=False)
    result = grouped.size().reset_index(name="row_count")
    if "__target__" in work:
        target_group = grouped["__target__"].agg(["sum", "mean"]).reset_index().rename(columns={"sum": "bad_count", "mean": "bad_rate"})
        result = result.merge(target_group, on=group_cols, how="left")
    result = result[result["row_count"] >= int(spec.get("min_group_size", 50))]
    result = result.sort_values("row_count", ascending=False).head(max_groups)
    records = result.to_dict(orient="records")
    return {"dimensions": columns, "rows": records, "suppressed_groups": max(0, int(len(grouped)) - len(records)), "spec": spec}
